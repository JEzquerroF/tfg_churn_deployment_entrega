"""
interpretation_dict.py — genera el diccionario de interpretación (Excel estático).

A diferencia del Excel operacional (que depende del análisis), este fichero es
ESTÁTICO: se genera desde config/archetypes.yaml + el catálogo de contramedidas,
y sirve como guía para que el cliente sepa leer predictions.csv y los códigos.

Hojas:
  - "Como leer la tabla": explica cada columna de predictions.csv + un ejemplo.
  - "Arquetipos": código + nombre + descripción + prioridad de negocio.
  - "Contramedidas": código + nombre + descripción + mecánica + disparador.
  - "Ejes de gusto": eje + qué mide + valores posibles.
  - "Codigos de origen": oof_stable / live_drift / live_new + explicación.

Uso:
    from reports.interpretation_dict import generate_interpretation_dict
    generate_interpretation_dict(repo_root, output_path)
"""

from __future__ import annotations

from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from reports.excel_generator import _auto_width, _format_header_row

TITLE_FONT = Font(bold=True, size=14, color="1F3A68")
SECTION_FONT = Font(bold=True, size=12, color="1F3A68")


# Qué mide cada eje (lenguaje claro) + valores posibles (contrato estable).
EJES_INFO = {
    "clase_favorita": (
        "La clase de personaje principal del jugador.",
        ["Caballero", "Asesino", "Campeón"],
    ),
    "estilo_build": (
        "Si el equipamiento del jugador es ofensivo (agresivo), defensivo (tanque) o equilibrado.",
        ["tanque", "equilibrado", "agresivo"],
    ),
    "monetizacion": (
        "El comportamiento de pago del jugador.",
        ["trial_no_convertido", "pagador", "no_pagador_ads", "no_pagador"],
    ),
    "perfil_enhance": (
        "Si el jugador invierte en mejorar (enhance) su equipo.",
        ["inversor", "no_inversor"],
    ),
    "perfil_coleccion": (
        "Si el jugador colecciona ítems o solo usa lo funcional.",
        ["coleccionista", "funcional"],
    ),
    "pvp_perfil": (
        "El comportamiento del jugador en combate PvP (solo disponible con actividad reciente).",
        ["pvp_frustrado", "pve_focus", "null"],
    ),
    "perfil_oro": (
        "El patrón de ahorro o gasto de moneda del jugador (solo disponible con actividad reciente).",
        ["acumulador", "gastador", "neutro", "null"],
    ),
}

# Explicación en lenguaje claro de los códigos de origen de la predicción.
SOURCE_CODES = {
    "oof_stable": (
        "Predicción estable: los datos del jugador no han cambiado respecto al "
        "entrenamiento, se usa la predicción histórica."
    ),
    "live_drift": (
        "El jugador ha cambiado de forma significativa desde el entrenamiento; "
        "se usa una predicción recalculada con sus datos actuales."
    ),
    "live_new": (
        "Jugador nuevo, no estaba en el entrenamiento; se usa una predicción "
        "calculada en vivo con sus datos."
    ),
}

# Columnas de predictions.csv explicadas.
COLUMN_DOCS = [
    ("user_id", "Identificador único del jugador."),
    ("churn_7d_final", "Probabilidad (0 a 1) de que el jugador abandone en los próximos 7 días."),
    ("churn_14d_final", "Probabilidad (0 a 1) de abandono en los próximos 14 días (horizonte de referencia)."),
    ("churn_30d_final", "Probabilidad (0 a 1) de abandono en los próximos 30 días."),
    ("archetype_name", "Arquetipo de jugador asignado (uno de los 6)."),
    ("contramedida_primaria_cod", "Código de la acción recomendada (CM01-CM12)."),
    ("contramedida_primaria_label", "Nombre corto de la acción recomendada."),
]


def _write_how_to_read(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Como leer la tabla"

    ws["A1"] = "CÓMO LEER predictions.csv"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = (
        "Cada fila de predictions.csv es un jugador. Estas son sus columnas:"
    )
    ws["A3"].font = Font(italic=True)

    row = 5
    for col_idx, header in enumerate(["Columna", "Qué significa"], start=1):
        ws.cell(row=row, column=col_idx, value=header)
    _format_header_row(ws, row)
    row += 1
    for name, doc in COLUMN_DOCS:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=doc)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Ejemplo de interpretación").font = SECTION_FONT
    row += 1
    ejemplo = (
        "Un jugador con churn_14d_final = 0.93, archetype_name = "
        "'Jugador Establecido Activo' y contramedida_primaria_cod = 'CM05' es un "
        "Jugador Establecido Activo con un 93% de probabilidad de abandonar en los "
        "próximos 14 días. Se le recomienda la contramedida CM05 (evento_enhance): "
        "un evento con más materiales o descuento en el coste de mejora de equipo."
    )
    cell = ws.cell(row=row, column=1, value=ejemplo)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 60

    _auto_width(ws)
    ws.column_dimensions["B"].width = 70


def _write_archetypes(wb: Workbook, archetypes_n1: dict) -> None:
    ws = wb.create_sheet(title="Arquetipos")
    ws["A1"] = "ARQUETIPOS DE JUGADOR"
    ws["A1"].font = TITLE_FONT

    row = 3
    for col_idx, header in enumerate(
        ["Código", "Nombre", "Descripción", "Prioridad de negocio"], start=1
    ):
        ws.cell(row=row, column=col_idx, value=header)
    _format_header_row(ws, row)
    row += 1
    for code, arch in archetypes_n1.items():
        ws.cell(row=row, column=1, value=int(code))
        ws.cell(row=row, column=2, value=arch.get("name", ""))
        ws.cell(row=row, column=3, value=arch.get("descripcion", "")).alignment = (
            Alignment(wrap_text=True, vertical="top")
        )
        ws.cell(row=row, column=4, value=arch.get("business_priority", ""))
        row += 1

    _auto_width(ws)
    ws.column_dimensions["C"].width = 70


def _write_countermeasures(wb: Workbook, contramedidas: list) -> None:
    ws = wb.create_sheet(title="Contramedidas")
    ws["A1"] = "CONTRAMEDIDAS (acciones recomendadas)"
    ws["A1"].font = TITLE_FONT

    row = 3
    for col_idx, header in enumerate(
        ["Código", "Nombre", "Descripción", "Mecánica", "Disparador"], start=1
    ):
        ws.cell(row=row, column=col_idx, value=header)
    _format_header_row(ws, row)
    row += 1
    for cm in sorted(contramedidas, key=lambda c: c.get("prioridad", 99)):
        ws.cell(row=row, column=1, value=cm.get("id", ""))
        ws.cell(row=row, column=2, value=cm.get("label", ""))
        ws.cell(row=row, column=3, value=cm.get("descripcion", "")).alignment = (
            Alignment(wrap_text=True, vertical="top")
        )
        ws.cell(row=row, column=4, value=cm.get("mecanica", "")).alignment = (
            Alignment(wrap_text=True, vertical="top")
        )
        ws.cell(row=row, column=5, value=" ".join(str(cm.get("disparador", "")).split()))
        row += 1

    _auto_width(ws)
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 45


def _write_axes(wb: Workbook) -> None:
    ws = wb.create_sheet(title="Ejes de gusto")
    ws["A1"] = "EJES DE PERFILADO DE GUSTOS"
    ws["A1"].font = TITLE_FONT

    row = 3
    for col_idx, header in enumerate(
        ["Eje", "Qué mide", "Valores posibles"], start=1
    ):
        ws.cell(row=row, column=col_idx, value=header)
    _format_header_row(ws, row)
    row += 1
    for eje, (que_mide, valores) in EJES_INFO.items():
        ws.cell(row=row, column=1, value=eje)
        ws.cell(row=row, column=2, value=que_mide).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.cell(row=row, column=3, value=", ".join(valores))
        row += 1

    _auto_width(ws)
    ws.column_dimensions["B"].width = 60


def _write_source_codes(wb: Workbook) -> None:
    ws = wb.create_sheet(title="Codigos de origen")
    ws["A1"] = "CÓDIGOS DE ORIGEN DE LA PREDICCIÓN"
    ws["A1"].font = TITLE_FONT

    row = 3
    for col_idx, header in enumerate(["Código", "Qué significa"], start=1):
        ws.cell(row=row, column=col_idx, value=header)
    _format_header_row(ws, row)
    row += 1
    for code, doc in SOURCE_CODES.items():
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=doc).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        row += 1

    _auto_width(ws)
    ws.column_dimensions["B"].width = 80


def generate_interpretation_dict(repo_root: Path, output_path: Path) -> Path:
    """
    Genera diccionario_interpretacion.xlsx (estático) desde config/archetypes.yaml.

    Args:
        repo_root: raíz del repo (para localizar config/archetypes.yaml).
        output_path: dónde escribir el .xlsx.

    Returns:
        Path al Excel generado.
    """
    cfg = yaml.safe_load((repo_root / "config" / "archetypes.yaml").read_text())
    archetypes_n1 = cfg.get("archetypes_n1", {})
    contramedidas = cfg.get("contramedidas", [])

    wb = Workbook()
    _write_how_to_read(wb)
    _write_archetypes(wb, archetypes_n1)
    _write_countermeasures(wb, contramedidas)
    _write_axes(wb)
    _write_source_codes(wb)

    wb.save(output_path)
    return output_path
