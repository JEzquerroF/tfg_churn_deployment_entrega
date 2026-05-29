"""
excel_generator.py — Genera el Excel operacional accionable.

Estructura del Excel:
  Hoja 1 (Resumen):     KPIs + tabla agregada por arquetipo.
  Hoja 2..N (por arquetipo): jugadores de cada arquetipo, ordenados por
                             riesgo (churn_30d_final desc).
  Hoja N+1 (Diccionario): mapping códigos → significados + contramedidas.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


HIGH_RISK_THRESHOLD = 0.65

HEADER_FILL = PatternFill(start_color="1F3A68", end_color="1F3A68", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
KPI_FILL = PatternFill(start_color="F0F4F9", end_color="F0F4F9", fill_type="solid")
KPI_FONT = Font(bold=True, size=14, color="1F3A68")


def _format_header_row(ws, row_num: int) -> None:
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _auto_width(ws) -> None:
    """Ajusta el ancho de las columnas según el contenido (capado a 50 chars)."""
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            # Saltar celdas en merged ranges (no tienen column_letter)
            if not hasattr(cell, "column_letter"):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        if col_letter is not None:
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)


def _section_title(ws, row: int, text: str) -> int:
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=12, color="1F3A68")
    return row + 1


def _table_header(ws, row: int, headers: list[str]) -> int:
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=header)
    _format_header_row(ws, row)
    return row + 1


def _pct(n: int, total: int) -> str:
    return f"{100 * n / total:.2f}%" if total else "0%"


def _write_summary_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    n_users: int,
    summary: Optional[dict],
    diccionario: Optional[dict],
) -> None:
    """Hoja Resumen legible: KPIs + alto riesgo por horizonte + distribuciones.

    Los agregados (arquetipos, contramedidas, cobertura, tier2) se toman de
    summary.json cuando está disponible; el alto riesgo por horizonte se calcula
    de las columnas churn_Xd_final de predictions.csv.
    """
    summary = summary or {}
    ws = wb.active
    ws.title = "Resumen"

    ws["A1"] = "RESUMEN EJECUTIVO — Predicción de churn"
    ws["A1"].font = Font(bold=True, size=16, color="1F3A68")
    ws.merge_cells("A1:E1")

    total = int(summary.get("n_users", n_users))
    ws["A3"] = "Total jugadores analizados"
    ws["B3"] = total
    ws.cell(row=3, column=1).font = Font(bold=True)
    ws.cell(row=3, column=2).font = KPI_FONT
    ws.cell(row=3, column=2).fill = KPI_FILL

    n_tier2 = summary.get("n_tier2")
    if n_tier2 is not None:
        ws["A4"] = "Usuarios con datos Tier 2 (actividad reciente)"
        ws["B4"] = int(n_tier2)
        ws.cell(row=4, column=1).font = Font(bold=True)
        ws.cell(row=4, column=2).font = KPI_FONT
        ws.cell(row=4, column=2).fill = KPI_FILL

    # --- Alto riesgo por horizonte (calculado de df) ---
    row = 6
    row = _section_title(ws, row, f"Alto riesgo de churn (probabilidad ≥ {HIGH_RISK_THRESHOLD:.2f})")
    row = _table_header(ws, row, ["Horizonte", "Jugadores", "% del total"])
    for horizon in ("7d", "14d", "30d"):
        col = f"churn_{horizon}_final"
        if col in df.columns:
            n_hr = int((df[col] >= HIGH_RISK_THRESHOLD).sum())
            ws.cell(row=row, column=1, value=horizon)
            ws.cell(row=row, column=2, value=n_hr)
            ws.cell(row=row, column=3, value=_pct(n_hr, total))
            row += 1

    # --- Distribución por arquetipo (de summary si está) ---
    arch_dist = summary.get("archetype_distribution")
    if not arch_dist and "archetype_name" in df.columns:
        arch_dist = df["archetype_name"].value_counts(dropna=False).to_dict()
    row += 1
    row = _section_title(ws, row, "Distribución por arquetipo")
    row = _table_header(ws, row, ["Arquetipo", "Jugadores", "% del total"])
    for name, n in sorted((arch_dist or {}).items(), key=lambda kv: -kv[1]):
        ws.cell(row=row, column=1, value=str(name))
        ws.cell(row=row, column=2, value=int(n))
        ws.cell(row=row, column=3, value=_pct(int(n), total))
        row += 1

    # --- Distribución de contramedidas (de summary + labels del diccionario) ---
    cm_dist = summary.get("countermeasure_distribution")
    if not cm_dist and "contramedida_primaria_cod" in df.columns:
        cm_dist = df["contramedida_primaria_cod"].value_counts(dropna=False).to_dict()
    cm_labels = {
        cod: cm.get("label", "")
        for cod, cm in (diccionario or {}).get("countermeasures", {}).items()
    }
    row += 1
    row = _section_title(ws, row, "Distribución de contramedidas")
    row = _table_header(ws, row, ["Código", "Acción", "Jugadores", "% del total"])
    for cod, n in sorted((cm_dist or {}).items(), key=lambda kv: -kv[1]):
        ws.cell(row=row, column=1, value=str(cod))
        ws.cell(row=row, column=2, value=cm_labels.get(cod, ""))
        ws.cell(row=row, column=3, value=int(n))
        ws.cell(row=row, column=4, value=_pct(int(n), total))
        row += 1

    coverage = summary.get("countermeasure_coverage_pct")
    if coverage is not None:
        ws.cell(row=row, column=1, value="Cobertura de contramedidas").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{coverage:.2f}%")
        ws.cell(row=row, column=2).font = KPI_FONT
        row += 1

    _auto_width(ws)


def _safe_sheet_name(name: str) -> str:
    """Limpia el nombre para cumplir restricciones de Excel."""
    cleaned = str(name)[:31]
    for ch in "/\\?*:[]":
        cleaned = cleaned.replace(ch, " ")
    return cleaned


def _write_archetype_sheets(wb: Workbook, df: pd.DataFrame) -> None:
    if "archetype_name" not in df.columns:
        return

    # Columnas operacionales: los 3 horizontes de churn + arquetipo + contramedida.
    cols_to_export = [
        "user_id",
        "archetype_name",
        "churn_7d_final",
        "churn_14d_final",
        "churn_30d_final",
        "contramedida_primaria_cod",
        "contramedida_primaria_label",
    ]
    cols_to_export = [c for c in cols_to_export if c in df.columns]

    # Cabeceras legibles (sin emojis).
    header_labels = {
        "user_id": "ID Jugador",
        "archetype_name": "Arquetipo",
        "churn_7d_final": "Prob. churn 7d",
        "churn_14d_final": "Prob. churn 14d",
        "churn_30d_final": "Prob. churn 30d",
        "contramedida_primaria_cod": "Contramedida",
        "contramedida_primaria_label": "Acción recomendada",
    }

    for archetype_name, group in df.groupby("archetype_name", dropna=False):
        if pd.isna(archetype_name):
            archetype_name = "Sin clasificar"

        sheet_name = _safe_sheet_name(archetype_name)
        ws = wb.create_sheet(title=sheet_name)

        sorted_group = group[cols_to_export].copy()
        if "churn_30d_final" in sorted_group.columns:
            sorted_group = sorted_group.sort_values(
                "churn_30d_final", ascending=False, na_position="last"
            )
        sorted_group = sorted_group.rename(columns=header_labels)

        for r_idx, row in enumerate(
            dataframe_to_rows(sorted_group, index=False, header=True), start=1
        ):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        _format_header_row(ws, 1)
        _auto_width(ws)


def generate_excel_report(
    predictions_path: Path,
    diccionario_path: Optional[Path],
    output_path: Path,
    summary_path: Optional[Path] = None,
) -> Path:
    """
    Genera el Excel operacional accionable.

    Args:
        predictions_path: ruta al predictions.csv.
        diccionario_path: ruta al diccionario.json (opcional).
        output_path: dónde escribir el .xlsx.
        summary_path: ruta al summary.json (opcional). Alimenta la hoja Resumen
            con cobertura de contramedidas y nº de usuarios Tier 2 (datos que no
            están en predictions.csv). Si falta, la hoja Resumen recalcula lo
            que puede desde predictions.csv.

    Returns:
        Path al Excel generado.
    """
    logger.info("Generando Excel: %s", output_path)

    df = pd.read_csv(predictions_path)
    n_users = len(df)

    diccionario: Optional[dict] = None
    if diccionario_path and Path(diccionario_path).exists():
        diccionario = json.loads(Path(diccionario_path).read_text())

    summary: Optional[dict] = None
    if summary_path and Path(summary_path).exists():
        summary = json.loads(Path(summary_path).read_text())

    wb = Workbook()
    _write_summary_sheet(wb, df, n_users, summary, diccionario)
    _write_archetype_sheets(wb, df)
    # La hoja "Diccionario" se eliminó del Excel operacional: la guía de
    # interpretación es ahora un fichero aparte (diccionario_interpretacion.xlsx,
    # generado por reports/interpretation_dict.py). El diccionario se sigue
    # leyendo para etiquetar las contramedidas en la hoja Resumen.

    wb.save(output_path)
    logger.info("Excel generado: %s (%.1f KB)", output_path, output_path.stat().st_size / 1024)
    return output_path
