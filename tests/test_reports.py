"""Tests de generación de informes PDF y Excel (Fase F.2)."""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from reports.excel_generator import generate_excel_report  # noqa: E402
from reports.interpretation_dict import generate_interpretation_dict  # noqa: E402
from reports.pdf_generator import generate_pdf_report  # noqa: E402


ARCHETYPE_NAMES = [
    "Recién Llegado Explorador",
    "Jugador Establecido Activo",
    "Hardcore End-Game",
    "Veterano Especializado",
    "Casual Dormido",
    "Veterano Inversor",
]


@pytest.fixture
def sample_outputs(tmp_path: Path) -> dict:
    """Outputs sintéticos mínimos para alimentar los generadores."""
    n = 100
    cm_codes = [f"CM{(i % 12) + 1:02d}" for i in range(n)]
    predictions = pd.DataFrame({
        "user_id": [f"user_{i:04d}" for i in range(n)],
        "archetype_n1": [i % 6 for i in range(n)],
        "archetype_name": [ARCHETYPE_NAMES[i % 6] for i in range(n)],
        "churn_7d_final": [0.5 + (i % 50) / 100 for i in range(n)],
        "churn_14d_final": [0.4 + (i % 50) / 100 for i in range(n)],
        "churn_30d_final": [0.3 + (i % 50) / 100 for i in range(n)],
        "churn_30d_source": ["live_new"] * n,
        "churn_30d_delta": [None] * n,
        "contramedida_primaria_cod": cm_codes,
        "contramedida_primaria_label": [f"accion_{c.lower()}" for c in cm_codes],
    })
    predictions_path = tmp_path / "predictions.csv"
    predictions.to_csv(predictions_path, index=False)

    diccionario = {
        "version": "2.0",
        "model_version": "v2_rf_L22_2026-05-19",
        "archetype_n1": {
            str(i): {
                "code": i,
                "name": ARCHETYPE_NAMES[i],
                "business_priority": "alta",
            }
            for i in range(6)
        },
        "ejes_gusto": {
            "clase_favorita": {"fuente": "char_class_main", "requiere_tier2": False,
                               "valores": ["Caballero", "Asesino", "Campeón"]},
            "pvp_perfil": {"fuente": "fights_pct_pvp, fights_pct_won",
                           "requiere_tier2": True,
                           "valores": ["pvp_frustrado", "pve_focus", "null"]},
        },
        "countermeasures": {
            f"CM{i:02d}": {
                "cod": f"CM{i:02d}",
                "label": f"accion_cm{i:02d}",
                "prioridad": i,
                "mecanica": "mecánica test",
                "disparador": "disparador test",
            }
            for i in range(1, 13)
        },
        "source_codes": {
            "oof_stable": {"description": "test", "recommended_action": "test"},
            "live_drift": {"description": "test", "recommended_action": "test"},
            "live_new": {"description": "test", "recommended_action": "test"},
        },
    }
    diccionario_path = tmp_path / "diccionario.json"
    diccionario_path.write_text(json.dumps(diccionario))

    summary = {
        "n_users": n,
        "churn_targets": ["churn_7d", "churn_14d", "churn_30d"],
        "drift_threshold": 0.10,
        "stage_6_status": "ok",
        "archetype_distribution": predictions["archetype_name"].value_counts().to_dict(),
        "countermeasure_distribution": predictions["contramedida_primaria_cod"].value_counts().to_dict(),
        "countermeasure_coverage_pct": 100.0,
        "n_tier2": 30,
    }
    summary_path = tmp_path / "summary.json"
    summary_path.write_text(json.dumps(summary))

    return {
        "predictions_path": predictions_path,
        "diccionario_path": diccionario_path,
        "summary_path": summary_path,
        "tmp_path": tmp_path,
    }


def test_pdf_generation(sample_outputs: dict) -> None:
    """El PDF se genera, existe y tiene un tamaño razonable."""
    pdf_path = sample_outputs["tmp_path"] / "test.pdf"
    result = generate_pdf_report(
        predictions_path=sample_outputs["predictions_path"],
        summary_path=None,
        output_path=pdf_path,
    )

    assert result.exists()
    size = result.stat().st_size
    assert size > 5_000, f"PDF demasiado pequeño ({size} bytes)"
    assert result.suffix == ".pdf"

    # Header PDF debe empezar por %PDF-
    with open(result, "rb") as f:
        assert f.read(5) == b"%PDF-"


def test_excel_generation(sample_outputs: dict) -> None:
    """El Excel se genera con las hojas esperadas."""
    from openpyxl import load_workbook

    xlsx_path = sample_outputs["tmp_path"] / "test.xlsx"
    result = generate_excel_report(
        predictions_path=sample_outputs["predictions_path"],
        diccionario_path=sample_outputs["diccionario_path"],
        output_path=xlsx_path,
        summary_path=sample_outputs["summary_path"],
    )

    assert result.exists()
    wb = load_workbook(result)
    sheet_names = wb.sheetnames

    # Debe haber: Resumen + 6 hojas de arquetipos = 7 hojas. La hoja
    # "Diccionario" se eliminó del Excel operacional (es un fichero aparte).
    assert "Resumen" in sheet_names
    assert "Diccionario" not in sheet_names, (
        "El Excel operacional ya NO debe tener hoja Diccionario"
    )
    assert len(sheet_names) >= 7

    # Sanity check: hoja Resumen tiene el título
    ws = wb["Resumen"]
    assert "RESUMEN" in str(ws["A1"].value).upper()

    # Resumen debe incluir las secciones nuevas (de summary.json)
    resumen_text = " ".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )
    assert "Tier 2" in resumen_text, "Resumen sin nº usuarios Tier 2"
    assert "Cobertura de contramedidas" in resumen_text
    assert "Distribución de contramedidas" in resumen_text
    assert "7d" in resumen_text and "14d" in resumen_text and "30d" in resumen_text

    # Las hojas de arquetipo deben tener cabeceras de los 3 horizontes
    arch_ws = wb[[s for s in sheet_names if s != "Resumen"][0]]
    headers = [c.value for c in arch_ws[1]]
    assert "Prob. churn 7d" in headers
    assert "Prob. churn 14d" in headers
    assert "Prob. churn 30d" in headers
    assert "Acción recomendada" in headers


def test_excel_without_diccionario(sample_outputs: dict) -> None:
    """El Excel también funciona si no se pasa diccionario."""
    from openpyxl import load_workbook

    xlsx_path = sample_outputs["tmp_path"] / "test_no_dic.xlsx"
    result = generate_excel_report(
        predictions_path=sample_outputs["predictions_path"],
        diccionario_path=None,
        output_path=xlsx_path,
    )

    assert result.exists()
    wb = load_workbook(result)
    assert "Resumen" in wb.sheetnames
    # Sin diccionario → no debe existir la hoja Diccionario
    assert "Diccionario" not in wb.sheetnames


def test_interpretation_dict_generation(tmp_path: Path) -> None:
    """El diccionario de interpretación estático se genera con sus 5 hojas."""
    from openpyxl import load_workbook

    out = tmp_path / "diccionario_interpretacion.xlsx"
    result = generate_interpretation_dict(ROOT, out)

    assert result.exists()
    wb = load_workbook(result)
    expected = {
        "Como leer la tabla", "Arquetipos", "Contramedidas",
        "Ejes de gusto", "Codigos de origen",
    }
    assert expected.issubset(set(wb.sheetnames)), (
        f"faltan hojas: {expected - set(wb.sheetnames)}"
    )

    # La hoja Contramedidas debe tener las 12 con descripción
    ws = wb["Contramedidas"]
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    for i in range(1, 13):
        assert f"CM{i:02d}" in text, f"falta CM{i:02d} en el diccionario"
    # La hoja Arquetipos debe incluir descripciones en lenguaje claro
    arch_text = " ".join(
        str(c.value) for row in wb["Arquetipos"].iter_rows() for c in row if c.value
    )
    assert "Casual Dormido" in arch_text
    assert "abandono" in arch_text.lower()
